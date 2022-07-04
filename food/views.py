import time
import pint
from django.http import HttpResponse

from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
import openpyxl
from openpyxl import Workbook, load_workbook
from food.models import Category, Ingredient, IngredientStock
from .forms import IngredientForm, StockForm

def food_main_view(request):
    categories = Category.objects.all()
    context = {}

    return render(request, 'food/food.html', context)

class IngredientiList(ListView):
    template_name = 'food/ingredienti.html'
    model = Ingredient
    context_object_name = 'ingredients'

    def get_queryset(self):
        return Ingredient.objects.all()


def add_ing(request):
    name= request.POST.get('ingname')
    category = Category.objects.get(id=1)
    ing= Ingredient.objects.create(name=name, category=category)
    # return template
    ingredients = Ingredient.objects.all()
    return render(request, 'food/partials/ingredient-list.html', {'ingredients':ingredients})



def ingredients_main_view(request):
    categories = Category.objects.all()
    ingredients = Ingredient.objects.all()

    context = { 'categories': categories, 'ingredients':ingredients}

    return render(request, 'food/ingredients.html', context)


def createIngredient(request):
    form = IngredientForm(request.POST or None)
    form2 = StockForm(request.POST or None)
    for field in form:
        print("Field Error:", field.name,  field.errors)
    for field in form2:
        print("Field Error:", field.name,  field.errors)

    if form.is_valid() and form2.is_valid():
        parent = form2.save(commit=False)
        parent.save()


        child = form.save(commit=False)

        child.stock = parent
        child.save()

        return redirect(child.get_absolute_url())
    else:
        print("not valid")
    return render(request, 'food/create_ing.html', {'form': form, 'form2': form2})





def check_ingname(request):
    ingname = request.POST.get('name')
    if Ingredient.objects.filter(name__iexact=ingname).exists():
        return HttpResponse("<div id='name-error' class='error'>This name exist</div>")
    else:
        return HttpResponse("<div id='name-error' class='success'> This name OK</div>")


def category_list_view(request):

    categories = Category.objects.all()

    context = { 'categories': categories}

    return render(request, "food/categories.html", context)

def category_detail_view(request, id):
    obj = get_object_or_404(Category, pk=id)
    ings = Ingredient.objects.filter(category=obj)

    categories = Category.objects.all()

    context = {"object": obj,
               'ings': ings, 'categories': categories}

    return render(request, "food/category_detail.html", context)


def createCategory(request):
    return render(request, 'food/food.html', {})

# INGREDIENTS
#


def ingredient_active_list(request):
    ingredients = Ingredient.objects.filter(active=True)
    categories = Category.objects.all()
    context = {'ingredients': ingredients, 'categories': categories}
    return render(request, 'food/ingredients.html', context)


def ingredient_make_active(request, id=None):
    obj = get_object_or_404(Ingredient, pk=id)
    obj.active = True
    obj.save()
    ingredients = Ingredient.objects.filter(active=True)

    context = {'ingredients': ingredients}

    return render(request, 'food/ingredients.html', context)

def ingredient_make_inactive(request, id=None):
    obj = get_object_or_404(Ingredient, pk=id)
    obj.active = False
    obj.save()
    ingredients = Ingredient.objects.filter(active=True)

    context = {'ingredients': ingredients}

    return render(request, 'food/ingredients.html', context)






def editIngredient(request, id=None):
    obj = get_object_or_404(Ingredient, id=id)
    obj2= obj. stock
    form = IngredientForm(request.POST or None, instance=obj)
    form2 = StockForm(request.POST or None, instance=obj2)
    if form.is_valid() and form2.is_valid():
        parent = form2.save(commit=False)
        parent.save()
        print('pa', parent.id, type(parent))

        child = form.save(commit=False)
        print('cc', child.id, type(child))
        child.stock = parent
        child.save()

        return redirect(child.get_absolute_url())
    return render(request, 'food/edit_ing.html', {'form': form, 'form2': form2})





def ingredient_detail_view(request, id=None):
    obj = get_object_or_404(Ingredient, pk=id)
    context = {
        "object": obj
    }

    return render(request, "food/ingredient_detail.html", context)


# SWERVICE FUNCTIONS

def createstock():
    ingredients = Ingredient.objects.all()
    for ing in ingredients:
        stock = IngredientStock()
        print('stock ', stock)
        stock.quantity_as_float = 0
        stock.min_quantity_as_float = 0
        stock.save()
        ing.stock = stock
        ing.save()
        print('i ', ing, ing.stock)


def impcat(request):
    wb = load_workbook('/home/matte/DEVELOPMENT/res/categories.xlsx')

    ws = wb.active
    for row in range(2, 25):
        id = ws['A' + str(row)].value
        name = ws['B' + str(row)].value
        c = Category()

        c.name = name

        # c.save()
    ureg = pint.UnitRegistry()
    print(ureg['meters'])

    return render(request, 'food/food.html', {})


def imping(request):
    wb = load_workbook('/home/matte/DEVELOPMENT/res/ingred.xlsx')

    ws = wb.active
    for row in range(2, 851):
        id = ws['A' + str(row)].value
        name = ws['B' + str(row)].value

        catid = ws['C' + str(row)].value
        am = ws['D' + str(row)].value
        cal = ws['E' + str(row)].value
        glu = ws['F' + str(row)].value
        pro = ws['G' + str(row)].value
        lip = ws['H' + str(row)].value

        cate = Category.objects.get(pk=catid)
        ing = Ingredient()
        ing.name = name
        ing.category = cate
        ing.amido = am
        ing.calorie = cal
        ing.glucidi = glu
        ing.proteine = pro
        ing.lipidi = lip

        time.sleep(0.01)

        print('saved {}'.format(ing))
    #  ing.save()
    return render(request, 'food/food.html', {})


